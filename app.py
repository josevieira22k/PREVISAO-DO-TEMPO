import tkinter as tk  # Módulo para criar interface gráfica
from tkinter import messagebox  # Para exibir mensagens de erro e sucesso
import requests  # Para realizar requisições HTTP (buscar dados na API do OpenWeatherMap)
import subprocess  # Para abrir e fechar o navegador Chrome
from datetime import datetime  # Para capturar data e hora atuais
from threading import Timer  # Para programar tarefas assíncronas
from openpyxl import Workbook, load_workbook  # Para criar e manipular arquivos Excel
from openpyxl.styles import Font, PatternFill, Border, Side  # Para estilizar células no Excel

# Substitua pela sua chave de API
API_KEY = 'bfbccb1c6bc4841a8f1d4e7632c1bafc'

# Função para buscar a previsão do tempo
def buscar_previsao():
    cidade = entrada_cidade.get()  # Captura o nome da cidade digitada pelo usuário
    if not cidade:  # Verifica se o campo está vazio
        messagebox.showerror("Erro", "Por favor, insira o nome de uma cidade.")
        return
    
    # URL para buscar previsão do tempo com base no nome da cidade
    url = f'http://api.openweathermap.org/data/2.5/weather?q={cidade}&appid={API_KEY}&units=metric'

    try:
        response = requests.get(url)  # Envia a requisição para a API
        response.raise_for_status()  # Verifica se a requisição foi bem-sucedida
        data = response.json()  # Converte a resposta para JSON
        temperatura = round(data['main']['temp'], 1)  # Formata a temperatura para 1 casa decimal
        umidade = data['main']['humidity']  # Captura a umidade

        # Abre a URL da previsão no navegador Chrome
        abrir_url(f'https://www.google.com/search?q=weather+{cidade}', cidade, temperatura, umidade)
        
        limpar_campos()  # Limpa o campo de entrada após a busca
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao obter os dados: {e}")

# Função para abrir o navegador com a URL da previsão
def abrir_url(url, cidade, temperatura, umidade):
    # Caminho para o navegador Chrome
    chrome_path = 'C:/Program Files/Google/Chrome/Application/chrome.exe'
    try:
        # Abre o navegador Chrome com a URL como um novo processo
        subprocess.Popen([chrome_path, url])
        
        # Programa para fechar o navegador e salvar dados após 5 segundos
        Timer(5.0, fechar_navegador, [cidade, temperatura, umidade]).start()
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao abrir o navegador: {e}")

# Função para fechar o navegador e salvar os dados no Excel
def fechar_navegador(cidade, temperatura, umidade):
    try:
        # Fecha todos os processos do Chrome abertos
        subprocess.call("taskkill /F /IM chrome.exe", shell=True)
        # Salva os dados no arquivo Excel
        salvar_dados(cidade, temperatura, umidade)
        # Exibe mensagem de sucesso com detalhes da previsão
        messagebox.showinfo("Sucesso", f"Dados salvos para {cidade}.\nTemperatura: {temperatura:.1f}°C\nUmidade: {umidade}%")
    except Exception as e:
        print(f"Erro ao fechar o navegador: {e}")

# Função para salvar os dados no arquivo Excel
def salvar_dados(cidade, temperatura, umidade):
    try:
        # Tenta carregar o arquivo Excel existente
        workbook = load_workbook('historico_clima.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        # Se o arquivo não existir, cria um novo e configura o cabeçalho
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Histórico Climático'
        headers = ['Data', 'Hora', 'Cidade', 'Temperatura (°C)', 'Umidade (%)']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

    # Captura data e hora atuais
    data_hora = datetime.now()
    data = data_hora.strftime('%Y-%m-%d')
    hora = data_hora.strftime('%H:%M:%S')
    # Adiciona nova linha com os dados da previsão
    nova_linha = [data, hora, cidade, temperatura, umidade]
    sheet.append(nova_linha)

    ultima_linha = sheet.max_row  # Captura o número da última linha

    # Estiliza a célula de umidade com cor baseada no valor
    cell_umidade = sheet.cell(row=ultima_linha, column=5)
    if umidade < 12:
        cell_umidade.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Vermelho
    elif 12 <= umidade < 20:
        cell_umidade.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Laranja
    elif 20 <= umidade <= 30:
        cell_umidade.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarelo
    else:
        cell_umidade.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Verde para umidade acima de 30%

    # Estilizando a célula da umidade
    cell_umidade.font = Font(bold=False)

    workbook.save('historico_clima.xlsx')  # Salva o arquivo Excel
    print(f'Dados salvos: {data}, {hora}, Cidade: {cidade}, Temp: {temperatura:.0f}°C, Umidade: {umidade}%')

# Função para limpar o campo de entrada da cidade após a busca
def limpar_campos():
    entrada_cidade.delete(0, tk.END)

# Configuração da interface gráfica com Tkinter
janela = tk.Tk()  # Cria a janela principal
janela.title("Previsão do Tempo")  # Define o título da janela

# Label para instrução do usuário
label_cidade = tk.Label(janela, text="Digite o nome da cidade para a previsão do tempo:")
label_cidade.pack(pady=5)  # Adiciona o label à janela

# Campo de entrada para o nome da cidade
entrada_cidade = tk.Entry(janela, width=30)
entrada_cidade.pack(pady=5)  # Adiciona o campo de entrada à janela

# Botão para buscar a previsão do tempo
botao_buscar = tk.Button(janela, text="Buscar previsão", command=buscar_previsao)
botao_buscar.pack(pady=10)  # Adiciona o botão à janela

# Inicia o loop principal da interface
janela.mainloop()
