import tkinter as tk
from tkinter import messagebox
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# Substitua pela sua chave de API
API_KEY = 'bfbccb1c6bc4841a8f1d4e7632c1bafc'
LAT = '-15.4525'  # Latitude de Planaltina de Goiás
LON = '-47.6136'  # Longitude de Planaltina de Goiás
URL = f'http://api.openweathermap.org/data/2.5/weather?lat={LAT}&lon={LON}&appid={API_KEY}&units=metric'

def buscar_previsao():
    cidade = entrada_cidade.get()
    if not cidade:
        messagebox.showerror("Erro", "Por favor, insira o nome de uma cidade.")
        return
    
    # Construindo a URL para buscar a previsão
    url = f'http://api.openweathermap.org/data/2.5/weather?q={cidade}&appid={API_KEY}&units=metric'

    try:
        response = requests.get(url)
        response.raise_for_status()  # Verifica se a requisição foi bem-sucedida
        data = response.json()
        temperatura = round(data['main']['temp'], 1)  # Formata a temperatura para 1 casa decimal
        umidade = data['main']['humidity']
        salvar_dados(cidade, temperatura, umidade)
        messagebox.showinfo("Sucesso", f"Dados salvos para {cidade}.\nTemperatura: {temperatura:.1f}°C\nUmidade: {umidade}%")
        limpar_campos()  # Limpa o campo após a busca
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao obter os dados: {e}")

def salvar_dados(cidade, temperatura, umidade):
    try:
        # Tentando carregar a planilha existente
        workbook = load_workbook('historico_clima.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        # Criando uma nova planilha se não existir
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Histórico Climático'
        # Criando cabeçalhos e destacando em negrito
        headers = ['Data', 'Hora', 'Cidade', 'Temperatura (°C)', 'Umidade (%)']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            # Estilizando o cabeçalho
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

    # Separando data e hora
    data_atual = datetime.now().strftime('%Y-%m-%d')
    hora_atual = datetime.now().strftime('%H:%M:%S')
    nova_linha = [data_atual, hora_atual, cidade, temperatura, umidade]
    sheet.append(nova_linha)

    # Estilizando a célula de data para destacar em negrito
    ultima_linha = sheet.max_row
    cell_data = sheet.cell(row=ultima_linha, column=1)
    cell_data.font = Font(bold=True)

    # Salvando o arquivo
    workbook.save('historico_clima.xlsx')
    print(f'Dados salvos: {data_atual}, {hora_atual}, Cidade: {cidade}, Temp: {temperatura:.0f}°C, Umidade: {umidade}%')

def limpar_campos():
    # Função para limpar o campo de entrada
    entrada_cidade.delete(0, tk.END)

# Configurando a interface com Tkinter
janela = tk.Tk()
janela.title("Previsão do Tempo")

# Label e campo de entrada para a cidade
label_cidade = tk.Label(janela, text="Digite o nome da cidade na qual gostaria da previçao do tempo:")
label_cidade.pack(pady=5)

entrada_cidade = tk.Entry(janela, width=30)
entrada_cidade.pack(pady=5)

# Botão para buscar previsão
botao_buscar = tk.Button(janela, text="Buscar previsão", command=buscar_previsao)
botao_buscar.pack(pady=10)

# Iniciando a interface
janela.mainloop()
